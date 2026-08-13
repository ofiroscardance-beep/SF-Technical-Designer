"""Render a task-planner model into a real, RTL Salesforce implementation backlog .xlsx.

Produces the three-sheet workbook the delivery team consumes:
  1. מקרא ורקע      — intro, epic colour legend, status legend, object API map
  2. משימות ליישום  — the backlog: one row per implementation task
  3. פערים עסקיים   — open business gaps to clarify with the client (GAP-N)

Every sheet is true RTL (sheet_view.rightToLeft). Salesforce API names stay
English inside the cells; the sheet direction handles the Hebrew layout.
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

_MAX_CELL = 32767  # Excel's hard per-cell character limit

_BASE_FONT = "Arial"
_HEADER_FILL = "D9D9D9"
_EPIC_PALETTE = [
    "FCE4D6", "FFF2CC", "E2EFDA", "DDEBF7", "EAD1DC",
    "D9E1F2", "FFE699", "C6E0B4", "F8CBAD", "BDD7EE",
]

_TOP_RIGHT = Alignment(horizontal="right", vertical="top", wrap_text=True)
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_THIN = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_TASK_HEADERS = [
    "#", "Epic", "מזהה משימה", "כותרת המשימה", "סוג עבודה",
    "תיאור ומטרת המשימה", "אובייקטים ושדות API מדויקים (Salesforce)",
    "פרומפט מדויק ליישום בפועל", "תלות (מזהי משימות)", "סטטוס / פתוח לבירור",
]
_TASK_WIDTHS = [5, 16, 12, 30, 20, 46, 40, 60, 18, 24, 34]
_EPIC_COL = _TASK_HEADERS.index("Epic") + 1
_STATUS_COL = _TASK_HEADERS.index("סטטוס / פתוח לבירור") + 1
_GAP_HEADERS = [
    "#", "נושא", "מה עלה / הסתירה שנמצאה",
    "השאלה המדויקת ללקוח העסקי", "השפעה אם לא ייפתר לפני הפיתוח",
]
_GAP_WIDTHS = [5, 28, 50, 55, 45]


def _clean(value: object) -> object:
    """Sanitise an LLM-supplied cell value: strip control chars openpyxl rejects
    (IllegalCharacterError) and clamp to Excel's per-cell limit. Non-strings pass
    through untouched (numbers, blanks)."""
    if not isinstance(value, str):
        return value
    return ILLEGAL_CHARACTERS_RE.sub("", value)[:_MAX_CELL]


def _epic_colors(epics: list[dict]) -> dict[str, str]:
    return {
        epic["id"]: epic.get("color") or _EPIC_PALETTE[i % len(_EPIC_PALETTE)]
        for i, epic in enumerate(epics)
        if epic.get("id")
    }


def _set_widths(ws, widths: list[int]) -> None:
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width


def _write_header(ws, headers: list[str]) -> None:
    for c, text in enumerate(headers, start=1):
        cell = ws.cell(1, c, text)
        cell.font = Font(name=_BASE_FONT, bold=True)
        cell.fill = PatternFill("solid", fgColor=_HEADER_FILL)
        cell.alignment = _HEADER_ALIGN
        cell.border = _BORDER


def _build_background(ws, model: dict, colors: dict[str, str]) -> None:
    ws.sheet_view.rightToLeft = True
    ws.column_dimensions["A"].width = 100
    bg = model.get("background", {})
    r = 1
    ws.cell(r, 1, _clean(model.get("title", "טבלת משימות ליישום"))).font = Font(
        name=_BASE_FONT, bold=True, size=14
    )
    r += 2
    for line in bg.get("intro", []):
        ws.cell(r, 1, _clean(line)).alignment = _TOP_RIGHT
        r += 1
    if bg.get("source"):
        r += 1
        ws.cell(r, 1, _clean(bg["source"])).alignment = _TOP_RIGHT
        r += 1
    r += 1
    ws.cell(r, 1, "מקרא צבעים (עמודה Epic):").font = Font(name=_BASE_FONT, bold=True)
    r += 1
    for epic in model.get("epics", []):
        epic_id = epic.get("id")
        if not epic_id:
            continue
        cell = ws.cell(r, 1, _clean(epic.get("title", epic_id)))
        cell.fill = PatternFill("solid", fgColor=colors[epic_id])
        cell.alignment = _TOP_RIGHT
        r += 1
    r += 1
    ws.cell(r, 1, "מקרא סטטוס (צבע עמודת הסטטוס):").font = Font(name=_BASE_FONT, bold=True)
    r += 1
    for item in model.get("status_legend", []):
        parts = [p for p in (item.get("label"), item.get("desc")) if p]
        cell = ws.cell(r, 1, _clean(" - ".join(parts)))
        color = item.get("color")
        if color:
            cell.fill = PatternFill("solid", fgColor=color)
        cell.alignment = _TOP_RIGHT
        r += 1
    r += 1
    ws.cell(r, 1, "מיפוי אובייקטים - שם עסקי מול API Name:").font = Font(
        name=_BASE_FONT, bold=True
    )
    r += 1
    for obj in model.get("object_map", []):
        note = f"  |  {obj['note']}" if obj.get("note") else ""
        line = f"{obj.get('business', '')} = {obj.get('api', '')}{note}"
        ws.cell(r, 1, _clean(line)).alignment = _TOP_RIGHT
        r += 1


def _build_tasks(ws, model: dict, colors: dict[str, str], status_colors: dict[str, str]) -> None:
    ws.sheet_view.rightToLeft = True
    reviewer = model.get("background", {}).get("reviewer_column", "הערות סוקר")
    _write_header(ws, [*_TASK_HEADERS, reviewer])
    _set_widths(ws, _TASK_WIDTHS)
    ws.freeze_panes = "A2"
    for r, task in enumerate(model.get("tasks", []), start=2):
        values = [
            task.get("num", r - 1), task.get("epic", ""), task.get("task_id", ""),
            task.get("title", ""), task.get("work_type", ""), task.get("description", ""),
            task.get("api_objects", "-"), task.get("impl_prompt", ""),
            task.get("dependencies", "-"), task.get("status", ""), "",
        ]
        for c, value in enumerate(values, start=1):
            cell = ws.cell(r, c, _clean(value))
            cell.alignment = _TOP_RIGHT
            cell.border = _BORDER
        epic_color = colors.get(task.get("epic"))
        if epic_color:
            ws.cell(r, _EPIC_COL).fill = PatternFill("solid", fgColor=epic_color)
        status_color = status_colors.get(task.get("status_level"))
        if status_color:
            ws.cell(r, _STATUS_COL).fill = PatternFill("solid", fgColor=status_color)


def _build_gaps(ws, model: dict) -> None:
    ws.sheet_view.rightToLeft = True
    _write_header(ws, _GAP_HEADERS)
    _set_widths(ws, _GAP_WIDTHS)
    ws.freeze_panes = "A2"
    for r, gap in enumerate(model.get("gaps", []), start=2):
        values = [
            gap.get("num", r - 1), gap.get("topic", ""), gap.get("found", ""),
            gap.get("question", ""), gap.get("impact", ""),
        ]
        for c, value in enumerate(values, start=1):
            cell = ws.cell(r, c, _clean(value))
            cell.alignment = _TOP_RIGHT
            cell.border = _BORDER


def build_workbook(model: dict, output_path: str) -> str:
    """Render the task model to a three-sheet RTL .xlsx and return the output path."""
    colors = _epic_colors(model.get("epics", []))
    status_colors = {
        item["level"]: item["color"]
        for item in model.get("status_legend", [])
        if item.get("level") and item.get("color")
    }
    wb = Workbook()
    _build_background(wb.active, model, colors)
    wb.active.title = "מקרא ורקע"
    _build_tasks(wb.create_sheet("משימות ליישום"), model, colors, status_colors)
    _build_gaps(wb.create_sheet("פערים עסקיים לבירור מול הלקוח"), model)
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def _main() -> None:
    import json
    import sys

    if len(sys.argv) != 3:
        raise SystemExit("usage: python -m src.excel_builder <tasks.json> <output.xlsx>")
    model = json.loads(Path(sys.argv[1]).read_text(encoding="utf-8"))
    print(build_workbook(model, sys.argv[2]))


if __name__ == "__main__":
    _main()
