"""Render an entity/field model into the RTL "שדות פורמט טכני" .xlsx workbook.

Reproduces the reference field-mapping format:
  1. מקרא ישויות (ERD) — every entity in the spec's ERD, its API name and status
  2. one sheet per entity — "1. Standard Fields" then "2. Custom Fields", plus
     optional picklist value tables

Every sheet is true RTL (sheet_view.rightToLeft). Salesforce API names stay
English and left-aligned (LTR reading order); Hebrew columns read right-to-left.
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

_MAX_CELL = 32767  # Excel's hard per-cell character limit
_MAX_SHEET_NAME = 31  # Excel's hard sheet-name limit
_INVALID_SHEET_CHARS = set(r"[]:*?/\'")

_BASE_FONT = "Calibri"
_SECTION_FILL = "1F3864"
_STD_HEADER_FILL = "C9DAF8"
_CUSTOM_HEADER_FILL = "D0E0E3"
_INDEX_HEADER_FILL = "37474F"
_WHITE = "FFFFFF"

_THIN = Side(style="thin", color="CCCCCC")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_RIGHT_TOP = Alignment(horizontal="right", vertical="top", wrap_text=True)
_LEFT_TOP_LTR = Alignment(horizontal="left", vertical="top", wrap_text=True, readingOrder=1)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

_ORIGIN_STYLES = {
    "standard": ("C6EFCE", "קיים OOTB"),
    "org": ("DDEBF7", "קיים בסביבה"),
    "new": ("FFE699", "להקמה"),
    "unverified": ("F4CCCC", "לאימות"),
}
_ENTITY_STATUS_FILLS = {
    "exists": "C6EFCE",
    "extend": "FFE699",
    "build": "FFE699",
    "define": "FFF2CC",
    "missing": "F4CCCC",
}
_TBD = "להגדיר"

_FIELD_HEADERS = [
    "#", "Source Field Name", "Display Name", "Salesforce Field", "API Name",
    "Data Type", "Mandatory", "Notes / Mapping Rules", "מקור", "Source URL",
    "הערות סוקר",
]
_FIELD_KEYS = [
    "num", "source_field", "display", "sf_field", "api",
    "type", "mandatory", "notes", "origin", "source_url",
]
_FIELD_WIDTHS = [6, 24, 24, 26, 28, 18, 12, 44, 16, 42, 26]
_LTR_FIELD_COLS = {2, 4, 5, 6, 10}
_ORIGIN_COL = _FIELD_HEADERS.index("מקור") + 1

_INDEX_HEADERS = [
    "#", "ישות (שם עסקי)", "API Name", "Standard / Custom", "תפקיד בפתרון",
    "קשרים ב-ERD", "סטטוס", "Source URL", "גיליון",
]
_INDEX_KEYS = [
    "name", "api", "kind", "purpose", "relationships", "status", "source_url",
]
_INDEX_WIDTHS = [6, 28, 26, 18, 46, 40, 20, 42, 26]
_LTR_INDEX_COLS = {3, 8, 9}


def _clean(value: object) -> object:
    """Sanitise an LLM-supplied cell value: strip control chars openpyxl rejects
    and clamp to Excel's per-cell limit."""
    if not isinstance(value, str):
        return value
    return ILLEGAL_CHARACTERS_RE.sub("", value)[:_MAX_CELL]


def _set_widths(ws, widths: list[float]) -> None:
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width


def _write(ws, row: int, col: int, value: object, *, fill: str | None = None,
           font: Font | None = None, align: Alignment | None = None,
           border: bool = True):
    cell = ws.cell(row, col, _clean(value))
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)
    if font:
        cell.font = font
    if align:
        cell.alignment = align
    if border:
        cell.border = _BORDER
    return cell


def _write_header_row(ws, row: int, headers: list[str], fill: str,
                      color: str = "000000") -> None:
    for c, text in enumerate(headers, start=1):
        _write(ws, row, c, text, fill=fill, align=_HEADER_ALIGN,
               font=Font(name=_BASE_FONT, bold=True, color=color))


def _write_section_title(ws, row: int, text: str, span: int) -> None:
    _write(ws, row, 1, text, fill=_SECTION_FILL, align=_RIGHT_TOP,
           font=Font(name=_BASE_FONT, bold=True, size=12, color=_WHITE))
    for c in range(2, span + 1):
        _write(ws, row, c, None, fill=_SECTION_FILL)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    ws.row_dimensions[row].height = 21


def _sheet_name(entity: dict, used: set[str]) -> str:
    raw = entity.get("sheet") or f"{entity.get('name', 'Entity')} ({entity.get('kind', '')})"
    name = "".join(ch for ch in raw if ch not in _INVALID_SHEET_CHARS).strip()[:_MAX_SHEET_NAME]
    name = name or "Entity"
    candidate, i = name, 2
    while candidate.lower() in used:
        suffix = f" ({i})"
        candidate = name[:_MAX_SHEET_NAME - len(suffix)] + suffix
        i += 1
    used.add(candidate.lower())
    return candidate


def _write_field_rows(ws, row: int, fields: list[dict]) -> int:
    for i, field in enumerate(fields, start=1):
        origin = field.get("origin", "unverified")
        origin_fill, origin_label = _ORIGIN_STYLES.get(origin, _ORIGIN_STYLES["unverified"])
        for c, key in enumerate(_FIELD_KEYS, start=1):
            value = field.get(key, "")
            if key == "num":
                value = field.get("num", f"{i}.0")
            elif key == "origin":
                value = origin_label
            align = _CENTER if c == 1 else (
                _LEFT_TOP_LTR if c in _LTR_FIELD_COLS else _RIGHT_TOP)
            _write(ws, row, c, value, align=align, font=Font(name=_BASE_FONT, size=10),
                   fill=origin_fill if c == _ORIGIN_COL else None)
        _write(ws, row, len(_FIELD_HEADERS), None, align=_RIGHT_TOP)
        row += 1
    return row


def _write_field_section(ws, row: int, title: str, fields: list[dict],
                         header_fill: str) -> int:
    _write_section_title(ws, row, title, len(_FIELD_HEADERS))
    _write_header_row(ws, row + 1, _FIELD_HEADERS, header_fill)
    row = _write_field_rows(ws, row + 2, fields)
    if not fields:
        _write(ws, row, 1, _TBD, align=_CENTER, font=Font(name=_BASE_FONT, size=10))
        row += 1
    return row + 1


def _write_picklists(ws, row: int, picklists: list[dict]) -> int:
    for table in picklists:
        headers = table.get("headers", ["Salesforce Value", "Display Name"])
        _write_section_title(ws, row, table.get("title", "Picklist Values"), len(headers))
        _write_header_row(ws, row + 1, headers, _CUSTOM_HEADER_FILL)
        row += 2
        for values in table.get("rows", []):
            for c, value in enumerate(values, start=1):
                align = _LEFT_TOP_LTR if c == 1 else _RIGHT_TOP
                _write(ws, row, c, value, align=align, font=Font(name=_BASE_FONT, size=10))
            row += 1
        row += 1
    return row


def _build_entity(ws, entity: dict) -> None:
    ws.sheet_view.rightToLeft = True
    _set_widths(ws, _FIELD_WIDTHS)
    name = entity.get("name", "")
    row = _write_field_section(
        ws, 1, f"1. Standard Fields – {name}",
        entity.get("standard_fields", []), _STD_HEADER_FILL)
    row = _write_field_section(
        ws, row, f"2. Custom Fields – {name}",
        entity.get("custom_fields", []), _CUSTOM_HEADER_FILL)
    _write_picklists(ws, row, entity.get("picklists", []))


def _build_index(ws, model: dict, sheet_names: list[str]) -> None:
    ws.sheet_view.rightToLeft = True
    _set_widths(ws, _INDEX_WIDTHS)
    overview = model.get("overview", {})
    _write(ws, 1, 1, model.get("title", "שדות פורמט טכני"), border=False,
           font=Font(name=_BASE_FONT, bold=True, size=14))
    row = 3
    for line in overview.get("intro", []):
        _write(ws, row, 1, line, align=_RIGHT_TOP, border=False,
               font=Font(name=_BASE_FONT, size=10))
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        row += 1
    if overview.get("source"):
        _write(ws, row, 1, overview["source"], align=_RIGHT_TOP, border=False,
               font=Font(name=_BASE_FONT, size=10, italic=True))
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        row += 1
    row += 1
    _write(ws, row, 1, "מקרא עמודת מקור:", border=False,
           font=Font(name=_BASE_FONT, bold=True))
    for c, (fill, label) in enumerate(_ORIGIN_STYLES.values(), start=2):
        _write(ws, row, c, label, fill=fill, align=_CENTER,
               font=Font(name=_BASE_FONT, size=10))
    row += 2
    _write_header_row(ws, row, _INDEX_HEADERS, _INDEX_HEADER_FILL, color=_WHITE)
    _write_index_rows(ws, row + 1, model.get("entities", []), sheet_names)
    ws.freeze_panes = f"A{row + 1}"


def _write_index_rows(ws, row: int, entities: list[dict], sheet_names: list[str]) -> None:
    for i, (entity, sheet) in enumerate(zip(entities, sheet_names), start=1):
        _write(ws, row, 1, i, align=_CENTER, font=Font(name=_BASE_FONT, size=10))
        for c, key in enumerate(_INDEX_KEYS, start=2):
            fill = _ENTITY_STATUS_FILLS.get(entity.get("status_level", "")) \
                if key == "status" else None
            align = _LEFT_TOP_LTR if c in _LTR_INDEX_COLS else _RIGHT_TOP
            _write(ws, row, c, entity.get(key, _TBD if key == "status" else ""),
                   fill=fill, align=align, font=Font(name=_BASE_FONT, size=10))
        _write(ws, row, len(_INDEX_HEADERS), sheet, align=_LEFT_TOP_LTR,
               font=Font(name=_BASE_FONT, size=10))
        row += 1


def build_workbook(model: dict, output_path: str) -> str:
    """Render the entity/field model to the RTL field-mapping .xlsx; return the path."""
    entities = model.get("entities", [])
    wb = Workbook()
    used: set[str] = {"מקרא ישויות"}
    sheet_names = [_sheet_name(entity, used) for entity in entities]
    for entity, sheet in zip(entities, sheet_names):
        _build_entity(wb.create_sheet(sheet), entity)
    _build_index(wb.active, model, sheet_names)
    wb.active.title = "מקרא ישויות"
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def _main() -> None:
    import json
    import sys

    if len(sys.argv) != 3:
        raise SystemExit("usage: python -m src.fields_builder <fields.json> <output.xlsx>")
    model = json.loads(Path(sys.argv[1]).read_text(encoding="utf-8"))
    print(build_workbook(model, sys.argv[2]))


if __name__ == "__main__":
    _main()
