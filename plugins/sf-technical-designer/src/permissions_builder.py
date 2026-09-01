"""Render a permissions model into the fixed four-sheet RTL permissions audit .xlsx.

Reproduces the "AUDIT — הרשאות רוחביות" layout exactly, with generic content:
  1-תפקידים בתהליך      — every role by process stage
  2-מטריצת אובייקטים    — role (rows) x object (columns) CRUD matrix
  3-קבוצות ושיתוף       — the sharing model, widest exposure tier first
  4- טכני (לארכיטקט)    — licenses, permission sets, groups, build status

Every sheet is true RTL (sheet_view.rightToLeft). Salesforce API names stay
English inside the cells; the sheet direction handles the Hebrew layout.
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

_MAX_CELL = 32767  # Excel's hard per-cell character limit

_BASE_FONT = "Calibri"
_DARK_HEADER = "37474F"
_SUB_HEADER = "607D8B"
_SHARING_HEADER = "1F3864"
_ZEBRA = "F2F4F8"
_WHITE = "FFFFFF"

_THIN = Side(style="thin", color="CCCCCC")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_CENTER = Alignment(horizontal="center")
_CENTER_WRAP = Alignment(horizontal="center", wrap_text=True)
_RIGHT_MIDDLE = Alignment(horizontal="right", vertical="center", wrap_text=True)
_RIGHT_TOP = Alignment(horizontal="right", vertical="top", wrap_text=True)
_MIDDLE_WRAP = Alignment(vertical="center", wrap_text=True)

_STAGE_PALETTE = [
    "E8F5E9", "E3F2FD", "FFF9C4", "FFF3E0", "FCE4EC", "F3E5F5", "ECEFF1",
]
_TIER_PALETTE = [
    "B71C1C", "E65100", "F9A825", "7CB342", "26A69A", "5C6BC0", "37474F",
]
_CHANNEL_PALETTE = ["FFE0B2", "C8E6C9", "90CAF9", "E1BEE7", "FFCCBC", "B2DFDB"]
_LICENSE_PALETTE = ["FFE0B2", "E3F2FD", "C8E6C9", "E1BEE7", "FFCCBC", "B2DFDB"]
_STATUS_COLORS = {
    "exists": "C8E6C9",
    "extend": "FFE0B2",
    "build": "FFE0B2",
    "define": "FFF9C4",
    "missing": "FFCDD2",
}
_TBD = "להגדיר"
_NONE_MARK = "—"

_MATRIX_LEGEND = (
    "מקרא:  C = יצירה (Create)  ·  R = צפייה (Read)  ·  U = עריכה (Update)  ·  "
    "D = מחיקה (Delete)  ·  — = אין  ·  CRUD = הכל"
)
_SHARING_SUBTITLE = (
    "הרשאה רוחבית (Permission Set) = מה מותר לעשות · "
    "שיתוף (למטה) = אילו רשומות רואים בפועל"
)
_TECH_HEADERS = [
    "תפקיד", "ROLE", "רישיון בסיס (User)", "Permission Set License",
    "Permission Set Group", "Permission Sets ישירים", "קבוצות ציבוריות", "סטטוס",
]
_TECH_WIDTHS = [22, 14, 18, 30, 21.86, 20.86, 34.86, 18]
_TECH_KEYS = [
    "role", "channel", "license", "ps_license", "psg",
    "permission_sets", "public_groups", "status",
]


def _clean(value: object) -> object:
    """Sanitise an LLM-supplied cell value: strip control chars openpyxl rejects
    and clamp to Excel's per-cell limit."""
    if not isinstance(value, str):
        return value
    return ILLEGAL_CHARACTERS_RE.sub("", value)[:_MAX_CELL]


def _set_widths(ws, widths: list[float]) -> None:
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width


def _write(ws, row: int, col: int, value: object, *, fill: str | None = None,
           font: Font | None = None, align: Alignment | None = None,
           border: bool = True):
    cell = ws.cell(row, col, _clean(value))
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill if len(fill) == 8 else f"FF{fill}")
    if font:
        cell.font = font
    if align:
        cell.alignment = align
    if border:
        cell.border = _BORDER
    return cell


def _write_header_row(ws, row: int, headers: list[str], fill: str) -> None:
    for c, text in enumerate(headers, start=1):
        _write(ws, row, c, text, fill=fill, align=_HEADER_ALIGN,
               font=Font(name=_BASE_FONT, bold=True, color=_WHITE))


def _tag_colors(values: list[str], palette: list[str]) -> dict[str, str]:
    """Assign a stable colour per distinct tag value, by order of appearance."""
    distinct = list(dict.fromkeys(v for v in values if v))
    return {v: palette[i % len(palette)] for i, v in enumerate(distinct)}


def _access_style(value: str) -> tuple[str, str]:
    """Background and font colour for one CRUD matrix cell."""
    text = (value or "").strip()
    if text == "CRUD":
        return "43A047", _WHITE
    if text in ("", _NONE_MARK, "-"):
        return "ECEFF1", "777777"
    if "U" in text or "D" in text:
        return "A5D6A7", "000000"
    if "C" in text or "R" in text:
        return "90CAF9", "000000"
    return "FFF9C4", "000000"


def _build_process(ws, model: dict) -> None:
    process = model.get("process", {})
    ws.sheet_view.rightToLeft = True
    _set_widths(ws, [20, 34, 70])
    _write(ws, 1, 1, process.get("title", "כל בעלי התפקידים לפי שלב בתהליך"),
           font=Font(name=_BASE_FONT, bold=True, size=14), border=False)
    if process.get("flow"):
        _write(ws, 2, 1, process["flow"], border=False,
               font=Font(name=_BASE_FONT, color=_SUB_HEADER))
        ws.merge_cells("A2:C2")
    _write_header_row(ws, 4, ["שלב בתהליך", "מה קורה", "התפקידים המעורבים"], _DARK_HEADER)
    for i, stage in enumerate(process.get("stages", [])):
        row = 5 + i
        color = stage.get("color") or _STAGE_PALETTE[i % len(_STAGE_PALETTE)]
        _write(ws, row, 1, stage.get("stage", ""), fill=color, align=_MIDDLE_WRAP,
               font=Font(name=_BASE_FONT, bold=True, size=12))
        for col, key in ((2, "what"), (3, "roles")):
            _write(ws, row, col, stage.get(key, ""), fill=color, align=_MIDDLE_WRAP,
                   font=Font(name=_BASE_FONT, size=11))
        ws.row_dimensions[row].height = 39.75


def _build_matrix(ws, model: dict) -> None:
    matrix = model.get("matrix", {})
    objects = model.get("objects", [])
    ws.sheet_view.rightToLeft = True
    _write(ws, 1, 1, matrix.get("title", "מטריצת גישה — תפקיד (ימין) × אובייקט (למעלה)"),
           font=Font(name=_BASE_FONT, bold=True, size=12), border=False)
    _write(ws, 2, 1, matrix.get("legend", _MATRIX_LEGEND), border=False,
           font=Font(name=_BASE_FONT, bold=True, color="B71C1C"))
    _write_header_row(ws, 3, ["תפקיד", *[o.get("label", "") for o in objects]], _DARK_HEADER)
    _write(ws, 4, 1, "(שם מערכת)", fill=_SUB_HEADER, align=_CENTER_WRAP,
           font=Font(name=_BASE_FONT, size=8, color=_WHITE))
    for c, obj in enumerate(objects, start=2):
        _write(ws, 4, c, obj.get("api", ""), fill=_SUB_HEADER, align=_CENTER_WRAP,
               font=Font(name=_BASE_FONT, size=8, color=_WHITE))
    for i, entry in enumerate(matrix.get("rows", [])):
        row = 5 + i
        _write(ws, row, 1, entry.get("role", ""), align=Alignment(horizontal="right"),
               font=Font(name=_BASE_FONT, bold=True))
        access = entry.get("access", {})
        for c, obj in enumerate(objects, start=2):
            value = access.get(obj.get("api")) or access.get(obj.get("label")) or _TBD
            fill, color = _access_style(value)
            _write(ws, row, c, value, fill=fill, align=_CENTER,
                   font=Font(name=_BASE_FONT, size=9, color=color))
    _set_widths(ws, [24, *[11] * len(objects)])
    ws.freeze_panes = "B5"


def _build_sharing(ws, model: dict) -> None:
    sharing = model.get("sharing", {})
    ws.sheet_view.rightToLeft = True
    _set_widths(ws, [26, 30, 36])
    _write(ws, 1, 1, sharing.get("title", "מודל השיתוף — כל תפקיד ורמת החשיפה שלו (מהרחב לצר)"),
           font=Font(name=_BASE_FONT, bold=True, size=14, color=_SHARING_HEADER), border=False)
    _write(ws, 2, 1, sharing.get("subtitle", _SHARING_SUBTITLE), border=False,
           font=Font(name=_BASE_FONT, size=10))
    ws.merge_cells("A2:C2")
    _write_header_row(ws, 4, ["תפקיד", "מנגנון השיתוף", "דוגמה / הערה"], _SHARING_HEADER)
    row = 5
    for i, tier in enumerate(sharing.get("tiers", [])):
        color = tier.get("color") or _TIER_PALETTE[i % len(_TIER_PALETTE)]
        for col in range(1, 4):
            _write(ws, row, col, tier.get("title", "") if col == 1 else None, fill=color,
                   font=Font(name=_BASE_FONT, bold=True, size=12, color=_WHITE))
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        ws.row_dimensions[row].height = 24
        row += 1
        for j, entry in enumerate(tier.get("rows", [])):
            zebra = _ZEBRA if j % 2 else _WHITE
            _write(ws, row, 1, entry.get("role", ""), fill=zebra, align=_RIGHT_MIDDLE,
                   font=Font(name=_BASE_FONT, bold=True, size=10, color=_SHARING_HEADER))
            for col, key in ((2, "mechanism"), (3, "note")):
                _write(ws, row, col, entry.get(key, ""), fill=zebra, align=_RIGHT_MIDDLE,
                       font=Font(name=_BASE_FONT, size=10))
            ws.row_dimensions[row].height = 21.75
            row += 1


def _build_technical(ws, model: dict) -> None:
    technical = model.get("technical", {})
    rows = technical.get("rows", [])
    ws.sheet_view.rightToLeft = True
    _set_widths(ws, _TECH_WIDTHS)
    _write(ws, 1, 1, technical.get("title", "נתונים טכניים (לארכיטקט)"),
           font=Font(name=_BASE_FONT, bold=True, size=12), border=False)
    _write_header_row(ws, 2, _TECH_HEADERS, _DARK_HEADER)
    channel_colors = _tag_colors([r.get("channel", "") for r in rows], _CHANNEL_PALETTE)
    license_colors = _tag_colors([r.get("license", "") for r in rows], _LICENSE_PALETTE)
    for i, entry in enumerate(rows):
        row = 3 + i
        fills = {
            2: channel_colors.get(entry.get("channel", "")),
            3: license_colors.get(entry.get("license", "")),
            8: _STATUS_COLORS.get(entry.get("status_level", ""), "FFF9C4"),
        }
        for c, key in enumerate(_TECH_KEYS, start=1):
            if c == 1:
                align, font = _RIGHT_TOP, Font(name=_BASE_FONT, bold=True)
            elif c in (2, 3):
                align, font = _CENTER_WRAP, Font(name=_BASE_FONT, size=9 if c == 3 else 11)
            else:
                align, font = _RIGHT_TOP, Font(name=_BASE_FONT, size=9 if c == 4 else 11)
            _write(ws, row, c, entry.get(key, _NONE_MARK), fill=fills.get(c),
                   align=align, font=font)
    ws.freeze_panes = "B3"


def build_workbook(model: dict, output_path: str) -> str:
    """Render the permissions model to the four-sheet RTL .xlsx; return the path."""
    wb = Workbook()
    _build_process(wb.active, model)
    wb.active.title = "1-תפקידים בתהליך"
    _build_matrix(wb.create_sheet("2-מטריצת אובייקטים"), model)
    _build_sharing(wb.create_sheet("3-קבוצות ושיתוף"), model)
    _build_technical(wb.create_sheet("4- טכני (לארכיטקט)"), model)
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def _main() -> None:
    import json
    import sys

    if len(sys.argv) != 3:
        raise SystemExit("usage: python -m src.permissions_builder <permissions.json> <output.xlsx>")
    model = json.loads(Path(sys.argv[1]).read_text(encoding="utf-8"))
    print(build_workbook(model, sys.argv[2]))


if __name__ == "__main__":
    _main()
